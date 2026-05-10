from aiogram import Bot, Dispatcher, F
from aiogram.enums import ParseMode
from aiogram.client.default import DefaultBotProperties
from aiogram.types import (
    Message,
    ReplyKeyboardMarkup,
    KeyboardButton,
    ReplyKeyboardRemove,
)

from aiogram.filters import CommandStart
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.context import FSMContext
from aiogram.fsm.storage.memory import MemoryStorage
import sqlite3

import asyncio
import os
from openpyxl import Workbook
from aiogram.types import FSInputFile
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.types import FSInputFile
from aiogram.types import CallbackQuery

# =========================
# CONFIG
# =========================

TOKEN = os.getenv("BOT_TOKEN")

GROUP_CHAT_ID = -1003947220682

COACH_TOPIC_ID = 2
PLAYER_TOPIC_ID = 4

bot = Bot(
    token=TOKEN,
    default=DefaultBotProperties(parse_mode=ParseMode.HTML)
)

dp = Dispatcher(storage=MemoryStorage())

conn = sqlite3.connect("/data/database.db")
cursor = conn.cursor()

cursor.execute("""
CREATE TABLE IF NOT EXISTS coaches (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    full_name TEXT,
    birth_year TEXT,
    phone TEXT,
    photo TEXT,
    telegram TEXT,
    district TEXT,
    mahalla TEXT,
    experience TEXT,
    has_group TEXT,
    players_count TEXT,
    has_field TEXT,
    extra TEXT
)
""")

cursor.execute("""
CREATE TABLE IF NOT EXISTS players (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    full_name TEXT,
    birth_year TEXT,
    district TEXT,
    mahalla TEXT,
    photo TEXT,
    parent_phone TEXT,
    extra TEXT
)
""")

conn.commit()
# =========================
# KEYBOARDS
# =========================

districts_keyboard = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="Andijon shahri")],
        [KeyboardButton(text="Andijon tumani"), KeyboardButton(text="Asaka")],
        [KeyboardButton(text="Baliqchi"), KeyboardButton(text="Bo‘z")],
        [KeyboardButton(text="Buloqboshi"), KeyboardButton(text="Izboskan")],
        [KeyboardButton(text="Jalaquduq"), KeyboardButton(text="Marhamat")],
        [KeyboardButton(text="Oltinko‘l"), KeyboardButton(text="Paxtaobod")],
        [KeyboardButton(text="Qo‘rg‘ontepa"), KeyboardButton(text="Shahrixon")],
        [KeyboardButton(text="Ulug‘nor"), KeyboardButton(text="Xo‘jaobod")],
    ],
    resize_keyboard=True
)

main_menu = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="🧑‍🏫 Murabbiy sifatida murojaat")],
        [KeyboardButton(text="👦 Bolani ro‘yxatdan o‘tkazish")],
        [KeyboardButton(text="📚 Batafsil ma’lumot")],
        [KeyboardButton(text="📞 Aloqa")],
    ],
    resize_keyboard=True
)

info_menu = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="⭐ Nima uchun Andijon FK?")],
        [KeyboardButton(text="👨‍👩‍👦 Ota-onalar uchun")],
        [KeyboardButton(text="🧑‍🏫 Murabbiylar uchun")],
        [KeyboardButton(text="🚀 Futbolchilar uchun")],
        [KeyboardButton(text="💻 Elektron tizim")],
        [KeyboardButton(text="🏆 Liga va turnirlar")],
        [KeyboardButton(text="ℹ️ Loyiha haqida")],
        [KeyboardButton(text="⬅️ Orqaga")],
    ],
    resize_keyboard=True
)

phone_button = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="📲 Raqamni yuborish", request_contact=True)]
    ],
    resize_keyboard=True,
    one_time_keyboard=True
)

ha_yoq = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="✅ Ha")],
        [KeyboardButton(text="❌ Yo‘q")]
    ],
    resize_keyboard=True
)

skip_button = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="⏭ O‘tkazib yuborish")]
    ],
    resize_keyboard=True
)


contract_button = InlineKeyboardMarkup(
    inline_keyboard=[

        [
            InlineKeyboardButton(
                text="✅ Tanishdim va davom etaman",
                callback_data="accept_contract"
            )
        ]
    ]
)

coach_contract_button = InlineKeyboardMarkup(
    inline_keyboard=[

        [
            InlineKeyboardButton(
                text="✅ Tanishdim va davom etaman",
                callback_data="accept_coach_contract"
            )
        ]
    ]
)
# =========================
# STATES
# =========================

class CoachForm(StatesGroup):
    full_name = State()
    birth_year = State()
    phone = State()
    photo = State()
    telegram = State()
    district = State()
    mahalla = State()
    experience = State()
    has_group = State()
    players_count = State()
    has_field = State()
    extra = State()

class PlayerForm(StatesGroup):
    full_name = State()
    birth_year = State()
    district = State()
    mahalla = State()
    photo = State()
    parent_phone = State()
    extra = State()

# =========================
# START
# =========================

@dp.message(CommandStart())
async def start(message: Message):

    text = (
        "<b>⚽ Andijon FK rasmiy futbol to‘garaklari platformasiga xush kelibsiz.</b>\n\n"
        "Endi Andijon viloyatining barcha hududlarida bolalar "
        "mahallaning o‘zida shug‘ullanib turib ham klubning "
        "rasmiy tarbiyalanuvchisi bo‘lish imkoniyatiga ega bo‘ladi.\n\n"
        "Kerakli bo‘limni tanlang 👇"
    )

    await message.answer(text, reply_markup=main_menu)

# =========================
# INFO MENU
# =========================

@dp.message(F.text == "📚 Batafsil ma’lumot")
async def info_section(message: Message):

    text = (
        "<b>📚 Batafsil ma’lumot</b>\n\n"
        "Kerakli bo‘limni tanlang 👇"
    )

    await message.answer(text, reply_markup=info_menu)

# =========================
# BACK BUTTON
# =========================

@dp.message(F.text == "⬅️ Orqaga")
async def back_to_main(message: Message):

    await message.answer(
        "Asosiy menyu 👇",
        reply_markup=main_menu
    )

# =========================
# WHY ANDIJON FK
# =========================

@dp.message(F.text == "⭐ Nima uchun Andijon FK?")
async def why_andijon(message: Message):

    text = (
        "<b>⚽ Nima uchun Andijon FK rasmiy to‘garaklari?</b>\n\n"
        "✅ Rasmiy klub tizimi\n"
        "✅ Elektron nazorat tizimi\n"
        "✅ Davomat va statistika\n"
        "✅ Liga va turnirlar\n"
        "✅ Ota-onalar uchun monitoring\n"
        "✅ Akademiya sari imkoniyat\n"
        "✅ Iqtidorli futbolchilar uchun grant\n"
        "✅ Professional futbolga yo‘l"
    )

    await message.answer(text)

# =========================
# PARENTS INFO
# =========================

@dp.message(F.text == "👨‍👩‍👦 Ota-onalar uchun")
async def parents_info(message: Message):

    text = (
        "<b>👨‍👩‍👦 Ota-onalar uchun</b>\n\n"
        "Farzandingiz:\n\n"
        "✅ rasmiy klub tizimida shug‘ullanadi\n"
        "✅ davomat nazoratida bo‘ladi\n"
        "✅ statistikasi yuritiladi\n"
        "✅ murabbiy baholari shakllantiriladi\n"
        "✅ liga va turnirlarda qatnashadi\n\n"
        "Ota-onalar uchun maxsus elektron monitoring tizimi ishlab chiqilmoqda."
    )

    await message.answer(text)

# =========================
# CONTACT
# =========================

@dp.message(F.text == "📞 Aloqa")
async def contacts(message: Message):

    text = (
        "<b>📞 Aloqa</b>\n\n"
        "📱 Telefon: +998884727772\n"
        "📍 Manzil: Andijon\n"
        "📲 Telegram: @khamilofff\n"
        "📸 Instagram: instagram.com/@malades.bola"
    )

    await message.answer(text)

# =========================
# COACH FLOW
# =========================

@dp.message(F.text == "🧑‍🏫 Murabbiy sifatida murojaat")
async def coach_start(message: Message, state: FSMContext):

    await state.set_state(CoachForm.full_name)

    await message.answer(
        "📋 1/9\n\n"
        "F.I.SH kiriting 👇",
        reply_markup=ReplyKeyboardRemove()
    )

@dp.message(CoachForm.full_name)
async def coach_name(message: Message, state: FSMContext):

    await state.update_data(full_name=message.text)

    await state.set_state(CoachForm.birth_year)

    await message.answer(
        "📋 2/9\n\n"
        "Tug‘ilgan yilingizni kiriting"
    )

@dp.message(CoachForm.birth_year)
async def coach_birth(message: Message, state: FSMContext):

    if not message.text.isdigit():
        await message.answer("Faqat yil kiriting")
        return

    await state.update_data(birth_year=message.text)

    await state.set_state(CoachForm.phone)

    await message.answer(
        "📋 3/9\n\n"
        "Telefon raqamingizni yuboring 👇",
        reply_markup=phone_button
    )

@dp.message(CoachForm.phone)
async def coach_phone(message: Message, state: FSMContext):

    phone = message.contact.phone_number if message.contact else message.text

    await state.update_data(phone=phone)

    contract_file = FSInputFile(
        "coach_contract.pdf"
    )

    await message.answer_document(
        contract_file,
        caption=(
            "📄 Andijon FK murabbiylik hamkorlik shartnomasi\n\n"
            "⚠️ Minimal hamkorlik muddati: 3 oy\n\n"
            "Iltimos shartnoma bilan tanishib chiqing."
        ),
        reply_markup=coach_contract_button
    )

@dp.callback_query(F.data == "accept_coach_contract")
async def accept_coach_contract(callback: CallbackQuery, state: FSMContext):

    await state.set_state(CoachForm.telegram)

    await callback.message.answer(
        "📋 4/9\n\n"
        "Telegram username (ixtiyoriy)\n"
        "Misol: @username",
        reply_markup=ReplyKeyboardRemove()
    )

    await callback.answer()


@dp.message(CoachForm.telegram)
async def coach_telegram(message: Message, state: FSMContext):

    await state.update_data(telegram=message.text)

    await state.set_state(CoachForm.district)

    await message.answer(
        "📋 5/9\n\n"
        "Hududingizni tanlang 👇",
        reply_markup=districts_keyboard
    )

@dp.message(CoachForm.district)
async def coach_district(message: Message, state: FSMContext):

    await state.update_data(district=message.text)

    await state.set_state(CoachForm.mahalla)

    await message.answer(
        "📋 6/9\n\n"
        "Mahallangiz nomini kiriting 👇",
        reply_markup=ReplyKeyboardRemove()
    )

@dp.message(CoachForm.mahalla)
async def coach_mahalla(message: Message, state: FSMContext):

    await state.update_data(mahalla=message.text)

    await state.set_state(CoachForm.experience)

    await message.answer(
        "📋 7/9\n\n"
        "Futbol yoki murabbiylik tajribangiz haqida yozing"
    )

@dp.message(CoachForm.experience)
async def coach_experience(message: Message, state: FSMContext):

    await state.update_data(experience=message.text)

    await state.set_state(CoachForm.has_group)

    await message.answer(
        "📋 8/9\n\n"
        "Hozir guruhingiz bormi?",
        reply_markup=ha_yoq
    )

@dp.message(CoachForm.has_group)
async def coach_group(message: Message, state: FSMContext):

    await state.update_data(has_group=message.text)

    if message.text == "✅ Ha":

        await state.set_state(CoachForm.players_count)

        await message.answer(
            "Nechta bola shug‘ullanadi?",
            reply_markup=ReplyKeyboardRemove()
        )

    else:

        await state.update_data(players_count="0")

        await state.set_state(CoachForm.has_field)

        await message.answer(
            "📋 9/9\n\n"
            "Mashg‘ulot uchun maydon bormi?",
            reply_markup=ha_yoq
        )

@dp.message(CoachForm.players_count)
async def coach_players(message: Message, state: FSMContext):

    await state.update_data(players_count=message.text)

    await state.set_state(CoachForm.has_field)

    await message.answer(
        "📋 9/9\n\n"
        "Mashg‘ulot uchun maydon bormi?",
        reply_markup=ha_yoq
    )

@dp.message(CoachForm.has_field)
async def coach_field(message: Message, state: FSMContext):

    await state.update_data(has_field=message.text)
    
    await state.set_state(CoachForm.photo)
    
    await message.answer(
        "📸 O‘zingizning rasmingizni yuboring",
        reply_markup=ReplyKeyboardRemove()
    )

@dp.message(CoachForm.photo)
async def coach_photo(message: Message, state: FSMContext):

    if not message.photo:

        await message.answer(
            "📸 Rasm yuboring"
        )

        return

    photo_id = message.photo[-1].file_id

    await state.update_data(photo=photo_id)

    await state.set_state(CoachForm.extra)

    await message.answer(
        "Qo‘shimcha ma’lumot yuborishingiz mumkin",
        reply_markup=skip_button
    )

@dp.message(CoachForm.extra)
async def coach_finish(message: Message, state: FSMContext):

    extra = message.text

    if extra == "⏭ O‘tkazib yuborish":
        extra = "Yo‘q"

    await state.update_data(extra=extra)

    data = await state.get_data()

    admin_text = (
        f"🔔 <b>Yangi murabbiy arizasi</b>\n\n"
        f"👤 Ism: {data['full_name']}\n"
        f"📅 Tug‘ilgan yil: {data['birth_year']}\n"
        f"📞 Telefon: {data['phone']}\n"
        f"📲 Telegram: {data['telegram']}\n"
        f"📍 Hudud: {data['district']} / {data['mahalla']}\n"
        f"⚽ Tajriba: {data['experience']}\n"
        f"👥 Guruh: {data['has_group']}\n"
        f"👦 Bolalar soni: {data['players_count']}\n"
        f"🏟 Maydon: {data['has_field']}\n"
        f"📝 Qo‘shimcha: {data['extra']}"
    )

    cursor.execute("""
    INSERT INTO coaches (
        full_name,
        birth_year,
        phone,
        telegram,
        district,
        mahalla,
        experience,
        has_group,
        players_count,
        has_field,
        extra,
        photo
    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        data['full_name'],
        data['birth_year'],
        data['phone'],
        data['telegram'],
        data['district'],
        data['mahalla'],
        data['experience'],
        data['has_group'],
        data['players_count'],
        data['has_field'],
        data['extra'],
        data['photo']
    ))
    
    conn.commit()
    
    await bot.send_photo(
        chat_id=GROUP_CHAT_ID,
        photo=data['photo'],
        caption=admin_text,
        message_thread_id=COACH_TOPIC_ID
    )

    await message.answer(
        "✅ Murojaatingiz qabul qilindi.\n\n"
        "Mutaxassislarimiz siz bilan bog‘lanadi.\n\n"
        "Mahalladan katta futbolga.",
        reply_markup=main_menu
    )

    await state.clear()

# =========================
# PLAYER FLOW
# =========================

@dp.message(F.text == "👦 Bolani ro‘yxatdan o‘tkazish")
async def player_start(message: Message, state: FSMContext):

    await state.set_state(PlayerForm.full_name)

    await message.answer(
        "📋 1/5\n\n"
        "Farzandingiz F.I.SH kiriting 👇",
        reply_markup=ReplyKeyboardRemove()
    )

@dp.message(PlayerForm.full_name)
async def player_name(message: Message, state: FSMContext):

    await state.update_data(full_name=message.text)

    await state.set_state(PlayerForm.birth_year)

    await message.answer(
        "📋 2/5\n\n"
        "Bolaning tug‘ilgan yilini kiriting"
    )

@dp.message(PlayerForm.birth_year)
async def player_birth(message: Message, state: FSMContext):

    if not message.text.isdigit():
        await message.answer("Faqat yil kiriting")
        return

    await state.update_data(birth_year=message.text)

    await state.set_state(PlayerForm.district)

    await message.answer(
        "📋 3/5\n\n"
        "Hududingizni tanlang 👇",
        reply_markup=districts_keyboard
    )

@dp.message(PlayerForm.district)
async def player_district(message: Message, state: FSMContext):

    await state.update_data(district=message.text)

    await state.set_state(PlayerForm.mahalla)

    await message.answer(
        "📋 4/5\n\n"
        "Mahallangiz nomini kiriting 👇",
        reply_markup=ReplyKeyboardRemove()
    )

@dp.message(PlayerForm.mahalla)
async def player_mahalla(message: Message, state: FSMContext):

    await state.update_data(mahalla=message.text)

    await state.set_state(PlayerForm.parent_phone)

    await message.answer(
        "📋 5/5\n\n"
        "Ota-ona telefon raqamini yuboring 👇",
        reply_markup=phone_button
    )

@dp.message(PlayerForm.parent_phone)
async def player_phone(message: Message, state: FSMContext):

    phone = message.contact.phone_number if message.contact else message.text

    await state.update_data(parent_phone=phone)

    contract_file = FSInputFile(
        "contract.pdf"
    )

    await message.answer_document(
        contract_file,
        caption=(
            "📄 3 tomonlama shartnoma namunasi\n\n"
            "⚠️ Minimal muddat: 3 oy\n\n"
            "Iltimos shartnoma bilan tanishib chiqing."
        ),
        reply_markup=contract_button
    )

@dp.callback_query(F.data == "accept_contract")
async def accept_contract(callback: CallbackQuery, state: FSMContext):

    await state.set_state(PlayerForm.photo)

    await callback.message.answer(
        "📸 Futbolchi rasmini yuboring",
        reply_markup=ReplyKeyboardRemove()
    )

    await callback.answer()

@dp.message(PlayerForm.photo)
async def player_photo(message: Message, state: FSMContext):

    if not message.photo:

        await message.answer(
            "📸 Rasm yuboring"
        )

        return

    photo_id = message.photo[-1].file_id

    await state.update_data(photo=photo_id)

    await state.set_state(PlayerForm.extra)

    await message.answer(
        "Qo‘shimcha ma’lumot yuborishingiz mumkin",
        reply_markup=skip_button
    )

@dp.message(PlayerForm.extra)
async def player_finish(message: Message, state: FSMContext):

    extra = message.text

    if extra == "⏭ O‘tkazib yuborish":
        extra = "Yo‘q"

    await state.update_data(extra=extra)

    data = await state.get_data()

    admin_text = (
        f"👦 <b>Yangi bola ro‘yxatdan o‘tdi</b>\n\n"
        f"👤 Ism: {data['full_name']}\n"
        f"📅 Tug‘ilgan yil: {data['birth_year']}\n"
        f"📍 Hudud: {data['district']} / {data['mahalla']}\n"
        f"📞 Ota-ona: {data['parent_phone']}\n"
        f"📝 Qo‘shimcha: {data['extra']}"
    )

    cursor.execute("""
    INSERT INTO players (
        full_name,
        birth_year,
        district,
        mahalla,
        photo,
        parent_phone,
        extra
    ) VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (
        data['full_name'],
        data['birth_year'],
        data['district'],
        data['mahalla'],
        data['parent_phone'],
        data['photo'],
        data['extra']
    ))
    
    conn.commit()

    await bot.send_photo(
        chat_id=GROUP_CHAT_ID,
        photo=data['photo'],
        caption=admin_text,
        message_thread_id=PLAYER_TOPIC_ID
    )

    await message.answer(
        "✅ So‘rovingiz qabul qilindi.\n\n"
        "Hududingizdagi rasmiy to‘garak faoliyati boshlanishi bilan siz bilan bog‘lanamiz.",
        reply_markup=main_menu
    )

    await state.clear()


@dp.message(F.text == "/stat")
async def statistics(message: Message):

    # Umumiy futbolchilar
    cursor.execute("SELECT COUNT(*) FROM players")
    players_count = cursor.fetchone()[0]

    # Umumiy murabbiylar
    cursor.execute("SELECT COUNT(*) FROM coaches")
    coaches_count = cursor.fetchone()[0]

    # Futbolchilar hududlar bo‘yicha
    cursor.execute("""
        SELECT district, COUNT(*)
        FROM players
        GROUP BY district
    """)

    players_by_district = cursor.fetchall()

    # Murabbiylar hududlar bo‘yicha
    cursor.execute("""
        SELECT district, COUNT(*)
        FROM coaches
        GROUP BY district
    """)

    coaches_by_district = cursor.fetchall()

    # Dictga o‘tkazamiz
    players_dict = {district: count for district, count in players_by_district}
    coaches_dict = {district: count for district, count in coaches_by_district}

    # Barcha districtlarni birlashtiramiz
    all_districts = set(players_dict.keys()) | set(coaches_dict.keys())

    district_text = ""

    for district in sorted(all_districts):

        players = players_dict.get(district, 0)
        coaches = coaches_dict.get(district, 0)

        district_text += (
            f"📍 <b>{district}</b>\n"
            f"👦 Futbolchilar: {players}\n"
            f"🧑‍🏫 Murabbiylar: {coaches}\n\n"
        )

    text = (
        f"📊 <b>Andijon FK Statistikasi</b>\n\n"
        f"👦 Jami futbolchilar: {players_count}\n"
        f"🧑‍🏫 Jami murabbiylar: {coaches_count}\n\n"
        f"{district_text}"
    )

    await message.answer(text)
# =========================
# RUN BOT
# =========================            

@dp.message(F.text == "/export_players")
async def export_players(message: Message):

    wb = Workbook()
    ws = wb.active

    ws.title = "Futbolchilar"

    # Header
    ws.append([
        "ID",
        "Ism",
        "Tug‘ilgan yil",
        "Tuman",
        "Mahalla",
        "Telefon",
        "Qo‘shimcha"
    ])

    cursor.execute("""
        SELECT
            id,
            full_name,
            birth_year,
            district,
            mahalla,
            parent_phone,
            extra
        FROM players
    """)

    players = cursor.fetchall()

    for player in players:
        ws.append(player)

    file_name = "players.xlsx"

    wb.save(file_name)

    await message.answer_document(
        FSInputFile(file_name),
        caption="👦 Futbolchilar bazasi"
    )

@dp.message(F.text == "/export_coaches")
async def export_coaches(message: Message):

    wb = Workbook()
    ws = wb.active

    ws.title = "Murabbiylar"

    ws.append([
        "ID",
        "Ism",
        "Tug‘ilgan yil",
        "Telefon",
        "Telegram",
        "Tuman",
        "Mahalla",
        "Tajriba",
        "Guruh",
        "Bolalar soni",
        "Maydon",
        "Qo‘shimcha"
    ])

    cursor.execute("""
        SELECT
            id,
            full_name,
            birth_year,
            phone,
            telegram,
            district,
            mahalla,
            experience,
            has_group,
            players_count,
            has_field,
            extra
        FROM coaches
    """)

    coaches = cursor.fetchall()

    for coach in coaches:
        ws.append(coach)

    file_name = "coaches.xlsx"

    wb.save(file_name)

    await message.answer_document(
        FSInputFile(file_name),
        caption="🧑‍🏫 Murabbiylar bazasi"
    )

# =========================
# RUN BOT
# =========================

@dp.message(F.text.startswith("/players "))
async def district_players(message: Message):

    district = message.text.replace("/players ", "").strip()

    cursor.execute("""
        SELECT id, full_name, birth_year, parent_phone
        FROM players
        WHERE district LIKE ?
        ORDER BY id DESC
    """, (f"%{district}%",))

    players = cursor.fetchall()

    if not players:
        await message.answer("❌ Futbolchilar topilmadi")
        return

    text = f"👦 <b>{district}</b> futbolchilari:\n\n"

    for player in players:

        text += (
            f"👤 {player[0]}\n"
            f"📅 {player[1]}\n"
            f"📞 {player[2]}\n\n"
        )

    await message.answer(text)

@dp.message(F.text.startswith("/coaches "))
async def district_coaches(message: Message):

    district = message.text.replace("/coaches ", "").strip()

    cursor.execute("""
        SELECT full_name, phone
        FROM coaches
        WHERE district LIKE ?
        ORDER BY id DESC
    """, (f"%{district}%",))

    coaches = cursor.fetchall()

    if not coaches:
        await message.answer("❌ Murabbiy topilmadi")
        return

    text = f"🧑‍🏫 <b>{district}</b> murabbiylari:\n\n"

    for coach in coaches:

        text += (
            f"👤 {coach[0]}\n"
            f"📞 {coach[1]}\n\n"
        )

    await message.answer(text)

@dp.message(F.text.startswith("/delete_player "))
async def delete_player(message: Message):

    player_id = message.text.replace(
        "/delete_player ",
        ""
    ).strip()

    cursor.execute(
        "DELETE FROM players WHERE id = ?",
        (player_id,)
    )

    conn.commit()

    await message.answer(
        f"🗑 Futbolchi o‘chirildi: {player_id}"
    )

@dp.message(F.text.startswith("/delete_coach "))
async def delete_coach(message: Message):

    coach_id = message.text.replace(
        "/delete_coach ",
        ""
    ).strip()

    cursor.execute(
        "DELETE FROM coaches WHERE id = ?",
        (coach_id,)
    )

    conn.commit()

    await message.answer(
        f"🗑 Murabbiy o‘chirildi: {coach_id}"
    )

admin_menu = InlineKeyboardMarkup(
    inline_keyboard=[

        [
            InlineKeyboardButton(
                text="📊 Statistika",
                callback_data="stats"
            )
        ],

        [
            InlineKeyboardButton(
                text="👦 Futbolchilar",
                callback_data="players_menu"
            ),

            InlineKeyboardButton(
                text="🧑‍🏫 Murabbiylar",
                callback_data="coaches_menu"
            )
        ],

        [
            InlineKeyboardButton(
                text="📤 Export futbolchilar",
                callback_data="export_players"
            )
        ],

        [
            InlineKeyboardButton(
                text="📤 Export murabbiylar",
                callback_data="export_coaches"
            )
        ]
    ]
)


coaches_inline = InlineKeyboardMarkup(
    inline_keyboard=[

        [
            InlineKeyboardButton(
                text="Andijon shahri",
                callback_data="coach_Andijon shahri"
            )
        ],

        [
            InlineKeyboardButton(
                text="Asaka",
                callback_data="coach_Asaka"
            ),

            InlineKeyboardButton(
                text="Shahrixon",
                callback_data="coach_Shahrixon"
            )
        ],

        [
            InlineKeyboardButton(
                text="Izboskan",
                callback_data="coach_Izboskan"
            ),

            InlineKeyboardButton(
                text="Marhamat",
                callback_data="coach_Marhamat"
            )
        ],

        [
            InlineKeyboardButton(
                text="⬅️ Orqaga",
                callback_data="back_admin"
            )
        ]
    ]
)

@dp.callback_query(F.data == "coaches_menu")
async def coaches_menu(callback):

    cursor.execute("""
        SELECT district, COUNT(*)
        FROM coaches
        GROUP BY district
        ORDER BY COUNT(*) DESC
    """)

    districts = cursor.fetchall()

    keyboard = []

    for district in districts:

        keyboard.append([
            InlineKeyboardButton(
                text=f"{district[0]} ({district[1]})",
                callback_data=f"coach_{district[0]}"
            )
        ])

    keyboard.append([
        InlineKeyboardButton(
            text="⬅️ Orqaga",
            callback_data="back_admin"
        )
    ])

    coaches_keyboard = InlineKeyboardMarkup(
        inline_keyboard=keyboard
    )

    await callback.message.edit_text(
        "🧑‍🏫 Murabbiylar hududlari",
        reply_markup=coaches_keyboard
    )

    await callback.answer()


@dp.callback_query(
    F.data.startswith("coach_") &
    ~F.data.startswith("coachdetail_")
)
async def district_coaches_callback(callback):

    district = callback.data.replace(
        "coach_",
        ""
    )

    cursor.execute("""
        SELECT
            id,
            full_name,
            phone,
            experience
        FROM coaches
        WHERE district LIKE ?
        ORDER BY id DESC
    """, (f"%{district}%",))

    coaches = cursor.fetchall()

    if not coaches:

        await callback.message.answer(
            "❌ Murabbiy topilmadi"
        )

        return

    keyboard = []

    for coach in coaches[:20]:

        keyboard.append([
            InlineKeyboardButton(
                text=f"{coach[1]}",
                callback_data=f"coachdetail_{coach[0]}"
            )
        ])

    keyboard.append([
        InlineKeyboardButton(
            text="⬅️ Orqaga",
            callback_data="coaches_menu"
        )
    ])

    coaches_keyboard = InlineKeyboardMarkup(
        inline_keyboard=keyboard
    )

    await callback.message.edit_text(
        f"🧑‍🏫 {district} murabbiylari",
        reply_markup=coaches_keyboard
    )

    await callback.answer()

@dp.callback_query(F.data.startswith("coachdetail_"))
async def coach_detail(callback):

    coach_id = callback.data.replace(
        "coachdetail_",
        ""
    )

    cursor.execute("""
        SELECT
            id,
            full_name,
            birth_year,
            phone,
            telegram,
            district,
            mahalla,
            experience,
            has_group,
            players_count,
            has_field,
            extra
        FROM coaches
        WHERE id = ?
    """, (coach_id,))

    coach = cursor.fetchone()

    if not coach:
        return

    text = (
        f"🧑‍🏫 <b>Murabbiy ma'lumotlari</b>\n\n"
        f"🆔 {coach[0]}\n"
        f"👤 {coach[1]}\n"
        f"📅 {coach[2]}\n"
        f"📞 {coach[3]}\n"
        f"📲 {coach[4]}\n"
        f"📍 {coach[5]}\n"
        f"🏠 {coach[6]}\n"
        f"⚽ {coach[7]}\n"
        f"👥 Guruh: {coach[8]}\n"
        f"👦 Bolalar: {coach[9]}\n"
        f"🏟 Maydon: {coach[10]}\n"
        f"📝 {coach[11]}"
    )

    keyboard = InlineKeyboardMarkup(
        inline_keyboard=[

            [
                InlineKeyboardButton(
                    text="🗑 O‘chirish",
                    callback_data=f"delete_coach_{coach[0]}"
                )
            ],

            [
                InlineKeyboardButton(
                    text="⬅️ Orqaga",
                    callback_data=f"coach_{coach[5]}"
                )
            ]
        ]
    )

    await callback.message.edit_text(
        text,
        reply_markup=keyboard
    )

    await callback.answer()

@dp.callback_query(F.data.startswith("delete_coach_"))
async def delete_coach_button(callback):

    coach_id = callback.data.replace(
        "delete_coach_",
        ""
    )

    cursor.execute(
        "DELETE FROM coaches WHERE id = ?",
        (coach_id,)
    )

    conn.commit()

    await callback.message.edit_text(
        "🗑 Murabbiy o‘chirildi"
    )

    await callback.answer()



@dp.callback_query(F.data == "players_menu")
async def players_menu(callback):

    cursor.execute("""
        SELECT district, COUNT(*)
        FROM players
        GROUP BY district
        ORDER BY COUNT(*) DESC
    """)

    districts = cursor.fetchall()

    keyboard = []

    for index, district in enumerate(districts):

        keyboard.append([
            InlineKeyboardButton(
                text=f"{district[0]} ({district[1]})",
                callback_data=f"players_{index}"
            )
        ])

    keyboard.append([
        InlineKeyboardButton(
            text="⬅️ Orqaga",
            callback_data="back_admin"
        )
    ])

    players_keyboard = InlineKeyboardMarkup(
        inline_keyboard=keyboard
    )

    # districtlarni memoryga saqlaymiz
    dp["players_districts"] = districts

    await callback.message.edit_text(
        "👦 Futbolchilar hududlari",
        reply_markup=players_keyboard
    )

    await callback.answer()
    

@dp.callback_query(F.data.startswith("players_"))
async def district_players_callback(callback):

    index = int(
        callback.data.replace("players_", "")
    )

    districts = dp["players_districts"]

    district = districts[index][0]

    cursor.execute("""
        SELECT id, full_name, birth_year, parent_phone
        FROM players
        WHERE district = ?
        ORDER BY id DESC
    """, (district,))

    players = cursor.fetchall()

    if not players:

        await callback.message.answer(
            "❌ Futbolchilar topilmadi"
        )

        return

    keyboard = []

    for player in players[:20]:

        keyboard.append([
            InlineKeyboardButton(
                text=f"{player[1]} ({player[2]})",
                callback_data=f"player_{player[0]}"
            )
        ])

    keyboard.append([
        InlineKeyboardButton(
            text="⬅️ Orqaga",
            callback_data="players_menu"
        )
    ])

    players_keyboard = InlineKeyboardMarkup(
        inline_keyboard=keyboard
    )

    await callback.message.edit_text(
        f"👦 {district} futbolchilari",
        reply_markup=players_keyboard
    )

    await callback.answer()

@dp.callback_query(F.data.startswith("player_"))
async def player_detail(callback):

    player_id = callback.data.replace(
        "player_",
        ""
    )

    cursor.execute("""
        SELECT
            id,
            full_name,
            birth_year,
            district,
            mahalla,
            parent_phone,
            extra,
            status
        FROM players
        WHERE id = ?
    """, (player_id,))

    player = cursor.fetchone()

    if not player:
        return

    text = (
        f"👦 <b>Futbolchi ma'lumotlari</b>\n\n"
        f"🆔 {player[0]}\n"
        f"👤 {player[1]}\n"
        f"📅 {player[2]}\n"
        f"📍 {player[3]}\n"
        f"🏠 {player[4]}\n"
        f"📞 {player[5]}\n"
        f"📝 {player[6]}\n"
        f"📌 Status: {player[7]}"
    )
    
    keyboard = InlineKeyboardMarkup(
        inline_keyboard=[
    
            [
                InlineKeyboardButton(
                    text="📞 Aloqa qilindi",
                    callback_data=f"status_called_{player[0]}"
                )
            ],
    
            [
                InlineKeyboardButton(
                    text="📄 Shartnoma",
                    callback_data=f"status_contract_{player[0]}"
                )
            ],
    
            [
                InlineKeyboardButton(
                    text="⚽ Trial",
                    callback_data=f"status_trial_{player[0]}"
                )
            ],
    
            [
                InlineKeyboardButton(
                    text="✅ Qabul qilindi",
                    callback_data=f"status_accepted_{player[0]}"
                )
            ],
    
            [
                InlineKeyboardButton(
                    text="❌ Bekor qilindi",
                    callback_data=f"status_cancel_{player[0]}"
                )
            ],
    
            [
                InlineKeyboardButton(
                    text="🗑 O‘chirish",
                    callback_data=f"delete_player_{player[0]}"
                )
            ],
    
            [
                InlineKeyboardButton(
                    text="⬅️ Orqaga",
                    callback_data="players_menu"
                )
            ]
        ]
    )

    await callback.message.edit_text(
        text,
        reply_markup=keyboard
    )

    await callback.answer()

@dp.callback_query(F.data.startswith("delete_player_"))
async def delete_player_button(callback):

    player_id = callback.data.replace(
        "delete_player_",
        ""
    )

    cursor.execute(
        "DELETE FROM players WHERE id = ?",
        (player_id,)
    )

    conn.commit()

    await callback.message.edit_text(
        "🗑 Futbolchi o‘chirildi"
    )

    await callback.answer()

@dp.callback_query(F.data == "back_admin")
async def back_admin(callback):

    await callback.message.edit_text(
        "⚙️ Admin panel",
        reply_markup=admin_menu
    )

    await callback.answer()


@dp.message(F.text == "/admin")
async def admin_panel(message: Message):

    await message.answer(
        "⚙️ Admin panel",
        reply_markup=admin_menu
    )

@dp.callback_query(F.data == "stats")
async def stats_callback(callback):

    cursor.execute("SELECT COUNT(*) FROM players")
    players_count = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM coaches")
    coaches_count = cursor.fetchone()[0]

    text = (
        f"📊 Statistikalar\n\n"
        f"👦 Futbolchilar: {players_count}\n"
        f"🧑‍🏫 Murabbiylar: {coaches_count}"
    )

    await callback.message.answer(text)

    await callback.answer()

@dp.callback_query(F.data == "export_players")
async def export_players_callback(callback):

    wb = Workbook()
    ws = wb.active

    ws.title = "Futbolchilar"

    ws.append([
        "ID",
        "Ism",
        "Tug‘ilgan yil",
        "Tuman",
        "Mahalla",
        "Telefon",
        "Qo‘shimcha"
    ])

    cursor.execute("""
        SELECT
            id,
            full_name,
            birth_year,
            district,
            mahalla,
            parent_phone,
            extra
        FROM players
    """)

    players = cursor.fetchall()

    for player in players:
        ws.append(player)

    file_name = "players.xlsx"

    wb.save(file_name)

    await callback.message.answer_document(
        FSInputFile(file_name),
        caption="👦 Futbolchilar bazasi"
    )

    await callback.answer()

@dp.callback_query(F.data == "export_coaches")
async def export_coaches_callback(callback):

    wb = Workbook()
    ws = wb.active

    ws.title = "Murabbiylar"

    ws.append([
        "ID",
        "Ism",
        "Tug‘ilgan yil",
        "Telefon",
        "Telegram",
        "Tuman",
        "Mahalla",
        "Tajriba",
        "Guruh",
        "Bolalar soni",
        "Maydon",
        "Qo‘shimcha"
    ])

    cursor.execute("""
        SELECT
            id,
            full_name,
            birth_year,
            phone,
            telegram,
            district,
            mahalla,
            experience,
            has_group,
            players_count,
            has_field,
            extra
        FROM coaches
    """)

    coaches = cursor.fetchall()

    for coach in coaches:
        ws.append(coach)

    file_name = "coaches.xlsx"

    wb.save(file_name)

    await callback.message.answer_document(
        FSInputFile(file_name),
        caption="🧑‍🏫 Murabbiylar bazasi"
    )

    await callback.answer()


@dp.callback_query(F.data.startswith("status_"))
async def update_player_status(callback):

    data = callback.data.split("_")

    action = data[1]
    player_id = data[2]

    statuses = {
        "called": "📞 Aloqa qilindi",
        "contract": "📄 Shartnoma",
        "trial": "⚽ Trial",
        "accepted": "✅ Qabul qilindi",
        "cancel": "❌ Bekor qilindi"
    }

    status = statuses.get(action)

    cursor.execute(
        "UPDATE players SET status = ? WHERE id = ?",
        (status, player_id)
    )

    conn.commit()

    await callback.answer(
        "✅ Status yangilandi"
    )
# =========================
# RUN BOT
# =========================

async def main():

    print("Bot ishga tushdi...")

    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())
